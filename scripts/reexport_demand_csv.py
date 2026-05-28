"""Re-export lulc_demand_results.xlsx -> lulc_demand_results.csv with proper numeric formatting.

The existing CSV was corrupted by an Excel save-as-CSV step that inserted thousand
separators into the Confidence factor and Initial rate columns (turning the European
decimal "1,772004996" into the integer-formatted "1,772,004,996"). This script reads
the XLSX directly and writes a CSV with dot-as-decimal that the existing R clean_numeric
function will parse correctly.

Column names match the production code's expectations (snake_case).
"""
import csv
import openpyxl
import sys

SRC_XLSX = "E:/nascent-lulcc-agg/inputs/lulc/future_demand/LULC_demand_results.xlsx"
DST_CSV = "E:/nascent-lulcc-agg/inputs/lulc/future_demand/lulc_demand_results_fixed.csv"

# Map XLSX headers -> production CSV column names
COLUMN_MAP = {
    "ID": "ID",
    "participant name": "participant_name",
    "LULC": "lulc_name",
    "Scenario": "scenario",
    "Driver": "driver",
    "Region": "region_name",
    "Slider value": "slider_value",
    "curve": "curve_type",
    "Confidence factor": "confidence_factor",
    "Initial rate": "initial_rate",
}

NUMERIC_STRING_COLUMNS = {"confidence_factor", "initial_rate"}
NUMERIC_COLUMNS = {"slider_value"}


def parse_european_decimal(value):
    """Parse "1,772004996" -> 1.772004996. Returns float or original value if parse fails."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if not s:
        return ""
    n_commas = s.count(",")
    if n_commas == 0:
        try:
            return float(s)
        except ValueError:
            return s
    if n_commas == 1:
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return s
    # n_commas > 1: malformed (e.g. "1,772,004,996" thousand-separator artifact).
    # Best-effort: strip all commas and try as integer, but warn.
    stripped = s.replace(",", "")
    try:
        n = float(stripped)
        print(f"WARNING: multi-comma value {s!r} interpreted as {n} (stripped commas)",
              file=sys.stderr)
        return n
    except ValueError:
        return s


def main():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb.active

    headers_xlsx = [cell.value for cell in ws[1]]
    missing = [h for h in headers_xlsx if h not in COLUMN_MAP]
    if missing:
        print(f"WARNING: unmapped XLSX columns: {missing}", file=sys.stderr)
    headers_csv = [COLUMN_MAP.get(h, h) for h in headers_xlsx]

    rows_written = 0
    sample_rows = []
    with open(DST_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers_csv)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
            out_row = []
            for header_csv, value in zip(headers_csv, row):
                if header_csv in NUMERIC_STRING_COLUMNS:
                    out_row.append(parse_european_decimal(value))
                elif header_csv in NUMERIC_COLUMNS:
                    if value is None or value == "":
                        out_row.append("")
                    else:
                        try:
                            out_row.append(float(value))
                        except (TypeError, ValueError):
                            out_row.append(value)
                else:
                    out_row.append(value if value is not None else "")
            writer.writerow(out_row)
            rows_written += 1
            if rows_written <= 3:
                sample_rows.append(out_row)

    print(f"Wrote {rows_written} rows to {DST_CSV}")
    print(f"Headers: {headers_csv}")
    for i, r in enumerate(sample_rows, 1):
        print(f"Row {i}: {r}")


if __name__ == "__main__":
    main()
